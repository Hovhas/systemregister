#!/usr/bin/env python3
"""
Import System.xlsx → Systemregistrets API.

Kör: python scripts/import_xlsx.py [--api-url URL] [--org-id UUID] [--dry-run] System.xlsx
"""

import argparse
import logging
import sys
from datetime import date
from pathlib import Path

import time

import httpx
import openpyxl

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Enum-mappningar (svenska källvärden → API enum-strängar)
# ---------------------------------------------------------------------------

LIFECYCLE_MAP: dict[str, str] = {
    "Under förvaltning": "i_drift",
    "Under aktiv vidareutveckling": "under_inforande",
    "Under avveckling": "under_avveckling",
    "": "planerad",
}

CRITICALITY_MAP: dict[str, str] = {
    "Helt verksamhetskritiskt": "kritisk",
    "Till betydande del verksamhetskritiskt": "hög",
    "Till betydande del ej verksamhetskritiskt": "låg",
    "Ej verksamhetskritiskt": "låg",
    "": "medel",
}

# Kolumn 50: Systemkategorier → API enum
# API-enum: verksamhetssystem | stödsystem | infrastruktur | plattform | iot
CATEGORY_MAP: dict[str, str] = {
    "Kärnverksamhetssystem": "verksamhetssystem",
    "Stöd- och verksamhetsgemensamma system": "stödsystem",
    "Stödsystem": "stödsystem",
    "Infrastruktur": "infrastruktur",
    "Plattform": "plattform",
    "IoT": "iot",
    "iot": "iot",
}

# ---------------------------------------------------------------------------
# Kolumnindex (0-baserat)
# ---------------------------------------------------------------------------

COL_NAME = 0
COL_ORG_PATH = 1
COL_IT_OWNERS = 2
COL_BIZ_OWNERS = 3
COL_ACTIVE_USERS = 4
COL_INCIDENTS = 5
COL_IT_ASSESSMENT = 7
COL_HOSTING = 8
COL_IT_DESCRIPTION = 13
COL_LIFECYCLE = 15
COL_DECOMMISSION_YEAR = 16
COL_EXTERNAL_CONNECTIONS = 18
COL_CONTRACT_END = 20
COL_DRIFT_BUDGET = 22
COL_DEV_BUDGET = 23
COL_LOCK_IN = 25
COL_CRITICALITY = 26
COL_OVERLAPPING = 27
COL_BUSINESS_AREAS = 28
COL_DRIFT_YEAR = 29
COL_BIZ_DESCRIPTION = 30
COL_TAGS = 42
COL_DEPENDS_ON = 43
COL_DEPENDS_FROM = 44
COL_HAS_INFOSEC_CLASS = 46
COL_HAS_RISK_ANALYSIS = 47
COL_CATEGORIES = 50

# ---------------------------------------------------------------------------
# Hjälpfunktioner
# ---------------------------------------------------------------------------


def cell_str(row: tuple, col: int) -> str:
    """Returnera cell-värde som trimmad sträng, eller tom sträng."""
    try:
        val = row[col].value
        if val is None:
            return ""
        return str(val).strip()
    except IndexError:
        return ""


def cell_int(row: tuple, col: int) -> int | None:
    """Returnera cell-värde som int, eller None."""
    try:
        val = row[col].value
        if val is None:
            return None
        return int(float(str(val).replace(" ", "").replace("\xa0", "")))
    except (ValueError, IndexError):
        return None


def cell_date(row: tuple, col: int) -> date | None:
    """Returnera cell-värde som date, eller None."""
    try:
        val = row[col].value
        if val is None:
            return None
        if isinstance(val, date):
            return val
        # Försök parse ISO-format
        s = str(val).strip()
        if s:
            return date.fromisoformat(s[:10])
        return None
    except (ValueError, IndexError):
        return None


def parse_org_path(path: str) -> list[str]:
    """Dela orgsökväg på ' / ' och returnera lista med trimade delar."""
    return [p.strip() for p in path.split("/") if p.strip()]


def map_lifecycle(raw: str) -> str:
    """Mappa råvärde till lifecycle enum. Okänt värde → 'planerad'."""
    result = LIFECYCLE_MAP.get(raw)
    if result is None:
        log.debug("Okänt lifecycle-värde '%s', använder 'planerad'", raw)
        return "planerad"
    return result


def map_criticality(raw: str) -> str:
    """Mappa råvärde till criticality enum. Okänt värde → 'medel'."""
    result = CRITICALITY_MAP.get(raw)
    if result is None:
        log.debug("Okänd kritikalitet '%s', använder 'medel'", raw)
        return "medel"
    return result


def map_category(raw_categories: str) -> tuple[str | None, list[str]]:
    """
    Mappa kommaseparerade kategorier.

    Returnerar (primär_api_kategori, omatchade_kategorier).
    Primär = första matchade. Omatchade läggs i extended_attributes.
    """
    if not raw_categories.strip():
        return None, []

    parts = [p.strip() for p in raw_categories.split(",") if p.strip()]
    matched: str | None = None
    unmatched: list[str] = []

    for part in parts:
        api_val = CATEGORY_MAP.get(part)
        if api_val is not None:
            if matched is None:
                matched = api_val
        else:
            unmatched.append(part)

    return matched, unmatched


def hosting_model(raw: str) -> str | None:
    """Normalisera driftform till on-premise/cloud/hybrid."""
    lower = raw.lower()
    if "intern" in lower or "on-premise" in lower or "on premise" in lower:
        return "on-premise"
    if "extern" in lower and "hybrid" not in lower:
        return "cloud"
    if "hybrid" in lower:
        return "hybrid"
    if raw:
        return raw[:50]
    return None


# ---------------------------------------------------------------------------
# API-klient
# ---------------------------------------------------------------------------


class ApiClient:
    def __init__(self, base_url: str, org_id: str | None, dry_run: bool):
        self.base_url = base_url.rstrip("/")
        self.dry_run = dry_run
        headers = {"Content-Type": "application/json"}
        if org_id:
            headers["X-Organization-Id"] = org_id
        self.client = httpx.Client(headers=headers, timeout=30)

    def _url(self, path: str) -> str:
        return f"{self.base_url}{path}"

    def _request(self, method: str, path: str, retries: int = 3, **kwargs):
        for attempt in range(retries):
            try:
                resp = self.client.request(method, self._url(path), **kwargs)
                resp.raise_for_status()
                return resp
            except (httpx.ReadError, httpx.ConnectError, httpx.RemoteProtocolError) as e:
                if attempt < retries - 1:
                    wait = 2 ** attempt
                    log.warning("Retry %d/%d after %s (wait %ds)", attempt + 1, retries, e, wait)
                    time.sleep(wait)
                else:
                    raise

    def get_organizations(self) -> list[dict]:
        return self._request("GET", "/organizations/").json()

    def create_organization(self, payload: dict) -> dict:
        if self.dry_run:
            log.info("[DRY-RUN] POST /organizations %s", payload)
            return {"id": "00000000-0000-0000-0000-000000000000", **payload}
        return self._request("POST", "/organizations/", json=payload).json()

    def get_systems(self, organization_id: str) -> list[dict]:
        data = self._request(
            "GET", "/systems/",
            params={"organization_id": organization_id, "limit": 200},
        ).json()
        if isinstance(data, dict) and "items" in data:
            return data["items"]
        return data

    def create_system(self, payload: dict) -> dict:
        if self.dry_run:
            log.info("[DRY-RUN] POST /systems %s", payload.get("name"))
            return {"id": "00000000-0000-0000-0000-000000000000", **payload}
        return self._request("POST", "/systems/", json=payload).json()

    def close(self):
        self.client.close()


# ---------------------------------------------------------------------------
# Organisations-hantering
# ---------------------------------------------------------------------------


def build_org_registry(api: ApiClient) -> dict[str, str]:
    """
    Hämta befintliga organisationer och bygg {namn: id}-dict.
    """
    orgs = api.get_organizations()
    return {o["name"]: o["id"] for o in orgs}


def ensure_org_hierarchy(
    path_parts: list[str],
    org_registry: dict[str, str],
    api: ApiClient,
    stats: dict,
) -> str | None:
    """
    Säkerställ att hela hierarkin finns. Returnera leaf-organisationens id.

    path_parts: ["Sundsvall kommunkoncern", "Sundsvall", "Kultur- och fritidsförvaltningen"]
    """
    parent_id: str | None = None

    for i, name in enumerate(path_parts):
        if name in org_registry:
            parent_id = org_registry[name]
            continue

        # Bestäm org_type
        if i == 0:
            # Toppnivå är kommunen/kommunkoncernen
            org_type = "kommun"
        else:
            org_type = "kommun"  # förvaltningar är del av kommunen

        payload: dict = {
            "name": name,
            "org_number": None,
            "org_type": org_type,
            "parent_org_id": parent_id,
        }

        try:
            created = api.create_organization(payload)
            org_id = created["id"]
            org_registry[name] = org_id
            parent_id = org_id
            stats["orgs_created"] += 1
            log.info("Skapade organisation: %s (id=%s)", name, org_id)
        except httpx.HTTPStatusError as exc:
            log.error("Fel vid skapande av org '%s': %s", name, exc.response.text)
            return None

    return parent_id


# ---------------------------------------------------------------------------
# System-hantering
# ---------------------------------------------------------------------------


def build_system_set(api: ApiClient, org_id: str) -> set[str]:
    """Returnera set av systemnamn för given org (för dubblettkontroll)."""
    try:
        systems = api.get_systems(org_id)
        return {s["name"] for s in systems}
    except Exception:
        return set()


def row_to_system_payload(
    row: tuple,
    organization_id: str,
) -> dict:
    """Konvertera en Excel-rad till API-payload för POST /systems."""

    name = cell_str(row, COL_NAME)
    raw_lifecycle = cell_str(row, COL_LIFECYCLE)
    raw_criticality = cell_str(row, COL_CRITICALITY)
    raw_categories = cell_str(row, COL_CATEGORIES)
    raw_hosting = cell_str(row, COL_HOSTING)
    raw_biz_description = cell_str(row, COL_BIZ_DESCRIPTION)
    raw_it_description = cell_str(row, COL_IT_DESCRIPTION)

    # Beskrivning: kombinera IT + verksamhet om båda finns
    description_parts = [p for p in [raw_it_description, raw_biz_description] if p]
    description = " | ".join(description_parts) if description_parts else name

    # Livscykel
    lifecycle_status = map_lifecycle(raw_lifecycle)

    # Kritikalitet
    criticality = map_criticality(raw_criticality)

    # Systemkategori
    primary_category, unmatched_categories = map_category(raw_categories)
    system_category = primary_category or "verksamhetssystem"

    # Driftmiljö
    hosting = hosting_model(raw_hosting)

    # Verksamhetsområden (kommaseparerad → ta första)
    business_areas_raw = cell_str(row, COL_BUSINESS_AREAS)
    business_area: str | None = None
    if business_areas_raw:
        parts = [p.strip() for p in business_areas_raw.split(",") if p.strip()]
        business_area = parts[0][:255] if parts else None

    # Driftår → deployment_date
    drift_year_raw = cell_str(row, COL_DRIFT_YEAR)
    deployment_date: str | None = None
    if drift_year_raw:
        try:
            year = int(float(drift_year_raw))
            if 1980 <= year <= 2100:
                deployment_date = f"{year}-01-01"
        except ValueError:
            pass

    # Avvecklat år → planned_decommission_date
    decommission_year_raw = cell_str(row, COL_DECOMMISSION_YEAR)
    planned_decommission_date: str | None = None
    if decommission_year_raw:
        try:
            year = int(float(decommission_year_raw))
            if 1980 <= year <= 2100:
                planned_decommission_date = f"{year}-12-31"
        except ValueError:
            pass

    # Avtalstid
    contract_end = cell_date(row, COL_CONTRACT_END)
    contract_end_str = contract_end.isoformat() if contract_end else None

    # Budgetar
    drift_budget = cell_int(row, COL_DRIFT_BUDGET)
    dev_budget = cell_int(row, COL_DEV_BUDGET)

    # Taggar
    tags_raw = cell_str(row, COL_TAGS)
    tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []

    # Beroenden (fritext)
    depends_on = cell_str(row, COL_DEPENDS_ON)
    depends_from = cell_str(row, COL_DEPENDS_FROM)

    # Externa kopplingar
    external_connections = cell_str(row, COL_EXTERNAL_CONNECTIONS)

    # Säkerhet/compliance
    has_infosec = cell_str(row, COL_HAS_INFOSEC_CLASS).lower()
    has_risk = cell_str(row, COL_HAS_RISK_ANALYSIS).lower()
    infosec_done = has_infosec in ("ja", "yes", "true", "1", "x")
    risk_done = has_risk in ("ja", "yes", "true", "1", "x")

    # Aktiva användare och incidenter
    active_users = cell_int(row, COL_ACTIVE_USERS)
    incidents = cell_int(row, COL_INCIDENTS)

    # IT-bedömning
    it_assessment = cell_str(row, COL_IT_ASSESSMENT)

    # Inlåsningseffekter
    lock_in = cell_str(row, COL_LOCK_IN)

    # Överlappande funktionalitet
    overlapping = cell_str(row, COL_OVERLAPPING)

    # Ansvariga (email)
    it_owners = cell_str(row, COL_IT_OWNERS)
    biz_owners = cell_str(row, COL_BIZ_OWNERS)

    # Bygg extended_attributes med allt som inte mappar till direkta fält
    extended: dict = {}
    if unmatched_categories:
        extended["extra_kategorier"] = unmatched_categories
    if raw_categories:
        extended["systemkategorier_raw"] = raw_categories
    if tags:
        extended["taggar"] = tags
    if depends_on:
        extended["beroende_till"] = depends_on
    if depends_from:
        extended["beroende_från"] = depends_from
    if external_connections:
        extended["externa_kopplingar"] = external_connections
    if drift_budget is not None:
        extended["driftbudget_sek"] = drift_budget
    if dev_budget is not None:
        extended["utvecklingsbudget_sek"] = dev_budget
    if active_users is not None:
        extended["aktiva_användare"] = active_users
    if incidents is not None:
        extended["incidenter_förra_året"] = incidents
    if it_assessment:
        extended["it_bedömning"] = it_assessment
    if lock_in:
        extended["inlåsningseffekter"] = lock_in
    if overlapping:
        extended["överlappande_funktionalitet"] = overlapping
    if it_owners:
        extended["systemansvariga_it"] = it_owners
    if biz_owners:
        extended["systemansvariga_verksamhet"] = biz_owners
    if business_areas_raw:
        extended["verksamhetsområden"] = business_areas_raw
    if contract_end_str:
        extended["avtalstid_löper_ut"] = contract_end_str
    if infosec_done:
        extended["infosäkerhetsklassning_genomförd"] = True
    if risk_done:
        extended["riskanalys_genomförd"] = True

    payload: dict = {
        "organization_id": organization_id,
        "name": name,
        "description": description,
        "system_category": system_category,
        "lifecycle_status": lifecycle_status,
        "criticality": criticality,
        "hosting_model": hosting,
        "business_area": business_area,
        "extended_attributes": extended if extended else None,
    }

    if deployment_date:
        payload["deployment_date"] = deployment_date
    if planned_decommission_date:
        payload["planned_decommission_date"] = planned_decommission_date

    return payload


# ---------------------------------------------------------------------------
# Huvud-import
# ---------------------------------------------------------------------------


def import_xlsx(
    xlsx_path: Path,
    sheet_name: str,
    api: ApiClient,
) -> dict:
    stats = {
        "rows_total": 0,
        "orgs_created": 0,
        "systems_created": 0,
        "systems_skipped": 0,
        "systems_error": 0,
    }

    log.info("Öppnar %s, sheet '%s'", xlsx_path, sheet_name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    if sheet_name not in wb.sheetnames:
        log.error("Sheet '%s' finns inte. Tillgängliga: %s", sheet_name, wb.sheetnames)
        sys.exit(1)

    ws = wb[sheet_name]

    # --- Fas 1: Bygg organisations-register ---
    log.info("Fas 1: Läser orgsökvägar och skapar organisationer...")
    org_registry = build_org_registry(api)
    log.info("Befintliga organisationer: %d", len(org_registry))

    # Insamla alla unika sökvägar (hoppa över header-rad 1)
    rows = list(ws.iter_rows(min_row=2))
    stats["rows_total"] = len(rows)

    org_paths_seen: set[str] = set()
    for row in rows:
        name = cell_str(row, COL_NAME)
        if not name:
            continue
        org_path = cell_str(row, COL_ORG_PATH)
        if org_path and org_path not in org_paths_seen:
            org_paths_seen.add(org_path)
            parts = parse_org_path(org_path)
            if parts:
                ensure_org_hierarchy(parts, org_registry, api, stats)

    log.info(
        "Fas 1 klar: %d orgar skapade, totalt %d i register",
        stats["orgs_created"],
        len(org_registry),
    )

    # --- Fas 2: Importera system ---
    log.info("Fas 2: Importerar system...")

    # Cache: org_id → set av systemnamn (för dubblettkontroll)
    existing_systems: dict[str, set[str]] = {}

    for i, row in enumerate(rows, start=2):
        name = cell_str(row, COL_NAME)
        if not name:
            continue

        org_path = cell_str(row, COL_ORG_PATH)
        parts = parse_org_path(org_path)

        # Bestäm organisation
        org_id: str | None = None
        if parts:
            # Leaf-org är sista delen
            leaf_name = parts[-1]
            org_id = org_registry.get(leaf_name)
            if not org_id:
                # Fallback: försök hitta någon del av hierarkin
                for part in reversed(parts):
                    if part in org_registry:
                        org_id = org_registry[part]
                        break

        if not org_id:
            log.warning("Rad %d: Kan inte hitta org för '%s' (sökväg: %s)", i, name, org_path)
            stats["systems_error"] += 1
            continue

        # Dubblettkontroll
        if org_id not in existing_systems:
            existing_systems[org_id] = build_system_set(api, org_id)

        if name in existing_systems[org_id]:
            log.debug("Rad %d: System '%s' finns redan i org %s — hoppar över", i, name, org_id)
            stats["systems_skipped"] += 1
            continue

        # Bygg payload
        try:
            payload = row_to_system_payload(row, org_id)
        except Exception as exc:
            log.error("Rad %d: Fel vid mappning av '%s': %s", i, name, exc)
            stats["systems_error"] += 1
            continue

        # Skapa system
        try:
            created = api.create_system(payload)
            existing_systems[org_id].add(name)
            stats["systems_created"] += 1
            if stats["systems_created"] % 50 == 0:
                log.info("Progress: %d system skapade...", stats["systems_created"])
        except httpx.HTTPStatusError as exc:
            log.error(
                "Rad %d: API-fel för '%s': %s — %s",
                i,
                name,
                exc.response.status_code,
                exc.response.text[:200],
            )
            stats["systems_error"] += 1
        except Exception as exc:
            log.error("Rad %d: Oväntat fel för '%s': %s", i, name, exc)
            stats["systems_error"] += 1

    wb.close()
    return stats


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Importera System.xlsx till systemregistrets API."
    )
    parser.add_argument(
        "xlsx",
        nargs="?",
        default="System.xlsx",
        help="Sökväg till Excel-filen (default: System.xlsx)",
    )
    parser.add_argument(
        "--api-url",
        default="https://systemregister.hovhas.se/api/v1",
        help="API base URL",
    )
    parser.add_argument(
        "--sheet",
        default="Sundsvall kommun",
        help="Sheet-namn i Excel-filen",
    )
    parser.add_argument(
        "--org-id",
        default=None,
        help="UUID att skicka i X-Organization-Id header",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Simulera utan att skriva till API",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Visa DEBUG-loggning",
    )
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        log.error("Filen %s hittades inte", xlsx_path)
        sys.exit(1)

    api = ApiClient(
        base_url=args.api_url,
        org_id=args.org_id,
        dry_run=args.dry_run,
    )

    try:
        stats = import_xlsx(xlsx_path, args.sheet, api)
    finally:
        api.close()

    log.info("=" * 50)
    log.info("Import klar!")
    log.info("  Rader totalt:        %d", stats["rows_total"])
    log.info("  Organisationer skapade: %d", stats["orgs_created"])
    log.info("  System skapade:      %d", stats["systems_created"])
    log.info("  System hoppade över: %d", stats["systems_skipped"])
    log.info("  Fel:                 %d", stats["systems_error"])
    log.info("=" * 50)

    if stats["systems_error"] > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
