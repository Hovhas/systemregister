from datetime import datetime, timezone
from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, Depends, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.rls import get_rls_db
from app.models import System
from app.schemas import (
    ComplianceGapResponse, NIS2ReportResponse,
    GDPRReportResponse, AIReportResponse,
    ClassificationReportResponse, LifecycleReportResponse,
)
from app.services.report_service import ReportService

router = APIRouter(prefix="/reports", tags=["Rapporter"])


@router.get("/nis2", response_model=NIS2ReportResponse)
async def nis2_report(
    organization_id: UUID | None = Query(None, description="Filtrera per organisation"),
    db: AsyncSession = Depends(get_rls_db),
):
    """NIS2-compliance-rapport med sammanfattning och systemlista."""
    systems = await ReportService.get_nis2_systems(db, organization_id)
    return ReportService.build_nis2_report(systems)


@router.get("/nis2.xlsx")
async def nis2_report_xlsx(db: AsyncSession = Depends(get_rls_db)):
    """NIS2-compliance-rapport som Excel-fil."""
    systems = await ReportService.get_nis2_systems(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "NIS2-system"
    ws.append(["Namn", "Organisation-ID", "NIS2-klass", "Kritikalitet", "Senaste riskbedömning", "Personuppgifter"])
    for s in systems:
        ws.append([
            s.name, str(s.organization_id),
            s.nis2_classification.value if s.nis2_classification else "",
            s.criticality.value,
            s.last_risk_assessment_date.isoformat() if s.last_risk_assessment_date else "",
            "Ja" if len(s.gdpr_treatments) > 0 else "Nej",
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=nis2-rapport.xlsx"},
    )


# --- HTML/PDF rendering ---

_REPORT_CSS = """
    body { font-family: Arial, sans-serif; font-size: 12px; margin: 2cm; color: #000; }
    h1 { font-size: 18px; margin-bottom: 4px; }
    h2 { font-size: 15px; margin-top: 24px; margin-bottom: 8px; }
    .meta { color: #555; margin-bottom: 16px; }
    .summary { background: #f5f5f5; padding: 12px; margin-bottom: 20px; border: 1px solid #ddd; }
    .summary h2 { font-size: 14px; margin: 0 0 8px 0; }
    .summary ul { margin: 0; padding-left: 18px; }
    table { width: 100%; border-collapse: collapse; margin-bottom: 24px; }
    th { background: #2c5282; color: #fff; text-align: left; padding: 6px 8px; }
    td { padding: 5px 8px; border-bottom: 1px solid #e2e8f0; vertical-align: top; }
    tr:nth-child(even) { background: #f7fafc; }
    .no-gaps { color: #276749; font-style: italic; }
    @media print { body { margin: 1cm; } .summary { break-inside: avoid; } tr { break-inside: avoid; } }
"""


def _render_nis2_html(systems: list[System]) -> str:
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    total = len(systems)
    without_classification = sum(1 for s in systems if s.nis2_classification is None)
    without_risk = sum(1 for s in systems if s.last_risk_assessment_date is None)
    rows = ""
    for s in systems:
        classification = s.nis2_classification.value if s.nis2_classification else "—"
        risk_date = s.last_risk_assessment_date.isoformat() if s.last_risk_assessment_date else "—"
        personal_data = "Ja" if len(s.gdpr_treatments) > 0 else "Nej"
        owners = ", ".join(o.name for o in s.owners) or "—"
        rows += f"<tr><td>{s.name}</td><td>{classification}</td><td>{s.criticality.value}</td><td>{risk_date}</td><td>{personal_data}</td><td>{owners}</td></tr>"
    return f"""<!DOCTYPE html><html lang="sv"><head><meta charset="UTF-8"><title>NIS2-rapport</title><style>{_REPORT_CSS}</style></head><body>
<h1>NIS2-compliance-rapport</h1><p class="meta">Sundsvalls kommunkoncern — Genererad: {generated_at}</p>
<div class="summary"><h2>Sammanfattning</h2><ul><li>Totalt NIS2-klassade system: <strong>{total}</strong></li><li>Utan NIS2-klassificering: <strong>{without_classification}</strong></li><li>Utan riskbedömning: <strong>{without_risk}</strong></li></ul></div>
<table><thead><tr><th>Namn</th><th>NIS2-klass</th><th>Kritikalitet</th><th>Senaste riskbedömning</th><th>Personuppgifter</th><th>Ägare</th></tr></thead><tbody>{rows}</tbody></table></body></html>"""


def _render_compliance_gap_html(gap_data: dict) -> str:
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    gaps = gap_data["gaps"]
    total_gaps = gap_data["summary"]["total_gaps"]

    def _sys_rows(items: list[dict]) -> str:
        if not items:
            return '<tr><td colspan="2" class="no-gaps">Inga gap hittades</td></tr>'
        return "".join(f"<tr><td>{i['name']}</td><td>{i['organization_id']}</td></tr>" for i in items)

    def _contract_rows(items: list[dict]) -> str:
        if not items:
            return '<tr><td colspan="3" class="no-gaps">Inga utgående avtal</td></tr>'
        return "".join(f"<tr><td>{c['supplier_name'] or '—'}</td><td>{c['system_id']}</td><td>{c['contract_end'] or '—'}</td></tr>" for c in items)

    sections = ""
    for title, key, cols in [
        ("System utan klassning", "missing_classification", "<th>Namn</th><th>Organisation-ID</th>"),
        ("System utan ägare", "missing_owner", "<th>Namn</th><th>Organisation-ID</th>"),
        ("Personuppgifter utan GDPR-behandling", "personal_data_without_gdpr", "<th>Namn</th><th>Organisation-ID</th>"),
        ("NIS2-system utan riskbedömning", "nis2_without_risk_assessment", "<th>Namn</th><th>Organisation-ID</th>"),
    ]:
        sections += f"<h2>{title}</h2><table><thead><tr>{cols}</tr></thead><tbody>{_sys_rows(gaps[key])}</tbody></table>"
    sections += f'<h2>Utgående avtal (inom 90 dagar)</h2><table><thead><tr><th>Leverantör</th><th>System-ID</th><th>Avtalsslutt</th></tr></thead><tbody>{_contract_rows(gaps["expiring_contracts"])}</tbody></table>'

    return f"""<!DOCTYPE html><html lang="sv"><head><meta charset="UTF-8"><title>Compliance-gap-rapport</title><style>{_REPORT_CSS}</style></head><body>
<h1>Compliance-gap-rapport</h1><p class="meta">Sundsvalls kommunkoncern — Genererad: {generated_at}</p>
<div class="summary"><h2>Sammanfattning</h2><ul><li>Totalt antal gap: <strong>{total_gaps}</strong></li></ul></div>{sections}</body></html>"""


@router.get("/nis2.html", response_class=HTMLResponse)
async def nis2_report_html(db: AsyncSession = Depends(get_rls_db)):
    systems = await ReportService.get_nis2_systems(db)
    return HTMLResponse(content=_render_nis2_html(systems))


@router.get("/nis2.pdf")
async def nis2_report_pdf(db: AsyncSession = Depends(get_rls_db)):
    from weasyprint import HTML as WeasyHTML
    systems = await ReportService.get_nis2_systems(db)
    pdf_bytes = WeasyHTML(string=_render_nis2_html(systems)).write_pdf()
    return StreamingResponse(
        BytesIO(pdf_bytes), media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=nis2-rapport.pdf"},
    )


@router.get("/compliance-gap", response_model=ComplianceGapResponse)
async def compliance_gap(db: AsyncSession = Depends(get_rls_db)):
    return await ReportService.get_compliance_gap_data(db)


@router.get("/compliance-gap.pdf")
async def compliance_gap_pdf(db: AsyncSession = Depends(get_rls_db)):
    from weasyprint import HTML as WeasyHTML
    gap_data = await ReportService.get_compliance_gap_data(db)
    pdf_bytes = WeasyHTML(string=_render_compliance_gap_html(gap_data)).write_pdf()
    return StreamingResponse(
        BytesIO(pdf_bytes), media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=compliance-gap-rapport.pdf"},
    )


# --- GDPR-rapport ---

@router.get("/gdpr", response_model=GDPRReportResponse)
async def gdpr_report(
    organization_id: UUID | None = Query(None, description="Filtrera per organisation"),
    db: AsyncSession = Depends(get_rls_db),
):
    """GDPR-rapport med sammanfattning och systemlista."""
    return await ReportService.get_gdpr_report(db, organization_id)


@router.get("/gdpr.xlsx")
async def gdpr_report_xlsx(
    organization_id: UUID | None = Query(None),
    db: AsyncSession = Depends(get_rls_db),
):
    """GDPR-rapport som Excel-fil."""
    data = await ReportService.get_gdpr_report(db, organization_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "GDPR-system"
    ws.append(["Namn", "Organisation-ID", "PuB-avtal", "DPIA", "Tredjelandsöverföring", "Saknar behandling", "Antal behandlingar"])
    for s in data["systems"]:
        ws.append([
            s["name"], s["organization_id"],
            "Ja" if s["has_pub"] else "Nej",
            "Ja" if s["has_dpia"] else "Nej",
            "Ja" if s["third_country_transfer"] else "Nej",
            "Ja" if s["missing_treatment"] else "Nej",
            s["treatment_count"],
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=gdpr-rapport.xlsx"},
    )


# --- AI-rapport ---

@router.get("/ai", response_model=AIReportResponse)
async def ai_report(
    organization_id: UUID | None = Query(None, description="Filtrera per organisation"),
    db: AsyncSession = Depends(get_rls_db),
):
    """AI-förordningsrapport med sammanfattning och systemlista."""
    return await ReportService.get_ai_report(db, organization_id)


@router.get("/ai.xlsx")
async def ai_report_xlsx(
    organization_id: UUID | None = Query(None),
    db: AsyncSession = Depends(get_rls_db),
):
    """AI-förordningsrapport som Excel-fil."""
    data = await ReportService.get_ai_report(db, organization_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "AI-system"
    ws.append(["Namn", "Organisation-ID", "Riskklass", "FRIA-status", "Transparens uppfylld", "Beskrivning"])
    for s in data["systems"]:
        ws.append([
            s["name"], s["organization_id"],
            s["ai_risk_class"] or "",
            s["fria_status"] or "",
            "Ja" if s["ai_transparency_fulfilled"] else "Nej",
            s["ai_usage_description"] or "",
        ])
    if data["modules"]:
        ws_mod = wb.create_sheet("AI-moduler")
        ws_mod.append(["Namn", "Organisation-ID", "Riskklass"])
        for m in data["modules"]:
            ws_mod.append([m["name"], m["organization_id"], m["ai_risk_class"] or ""])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ai-rapport.xlsx"},
    )


# --- Klassningsstatus-rapport ---

@router.get("/classification-status", response_model=ClassificationReportResponse)
async def classification_status_report(
    organization_id: UUID | None = Query(None, description="Filtrera per organisation"),
    db: AsyncSession = Depends(get_rls_db),
):
    """Klassningsstatusrapport med sammanfattning och systemlista."""
    return await ReportService.get_classification_report(db, organization_id)


@router.get("/classification-status.xlsx")
async def classification_status_report_xlsx(
    organization_id: UUID | None = Query(None),
    db: AsyncSession = Depends(get_rls_db),
):
    """Klassningsstatusrapport som Excel-fil."""
    data = await ReportService.get_classification_report(db, organization_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Klassningsstatus"
    ws.append(["Namn", "Organisation-ID", "Har klassning", "Senaste klassningsdatum", "Utgången"])
    for s in data["systems"]:
        ws.append([
            s["name"], s["organization_id"],
            "Ja" if s["has_classification"] else "Nej",
            s["most_recent_date"] or "",
            "Ja" if s["is_expired"] else "Nej",
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=klassningsstatus-rapport.xlsx"},
    )


# --- Livscykel-rapport ---

@router.get("/lifecycle", response_model=LifecycleReportResponse)
async def lifecycle_report(
    organization_id: UUID | None = Query(None, description="Filtrera per organisation"),
    db: AsyncSession = Depends(get_rls_db),
):
    """Livscykelrapport med utgående avtal, end-of-support och avveckling."""
    return await ReportService.get_lifecycle_report(db, organization_id)


@router.get("/lifecycle.xlsx")
async def lifecycle_report_xlsx(
    organization_id: UUID | None = Query(None),
    db: AsyncSession = Depends(get_rls_db),
):
    """Livscykelrapport som Excel-fil."""
    data = await ReportService.get_lifecycle_report(db, organization_id)
    wb = Workbook()

    ws_contracts = wb.active
    ws_contracts.title = "Utgående avtal"
    ws_contracts.append(["Leverantör", "System-ID", "Slutdatum", "Dagar kvar"])
    for c in data["contracts"]:
        ws_contracts.append([c["supplier_name"], c["system_id"], c["contract_end"], c["days_remaining"]])

    ws_eos = wb.create_sheet("End-of-support")
    ws_eos.append(["Namn", "Organisation-ID", "Slutdatum support", "Dagar kvar"])
    for s in data["end_of_support_systems"]:
        ws_eos.append([s["name"], s["organization_id"], s.get("end_of_support_date", ""), s.get("days_remaining", "")])

    ws_decomm = wb.create_sheet("Under avveckling")
    ws_decomm.append(["Namn", "Organisation-ID", "Planerat avvecklingsdatum"])
    for s in data["decommissioning_systems"]:
        ws_decomm.append([s["name"], s["organization_id"], s.get("planned_decommission_date", "")])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=livscykel-rapport.xlsx"},
    )
