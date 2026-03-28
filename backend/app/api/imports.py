import csv
import io
import json
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.rls import get_rls_db
from app.models import System
from app.models.enums import OwnerRole
from app.models.models import Organization, SystemClassification, SystemOwner
from app.schemas import SystemCreate

router = APIRouter(prefix="/import", tags=["Import"])

SUPPORTED_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/csv",
    "application/csv",
    "application/json",
    "text/plain",  # Vissa klienter skickar CSV som text/plain
}


def _detect_format(filename: str, content_type: str | None) -> str:
    """Detektera filformat utifrån filnamn och content-type."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return "excel"
    if name.endswith(".csv"):
        return "csv"
    if name.endswith(".json"):
        return "json"
    # Fallback på content-type
    ct = (content_type or "").lower()
    if "spreadsheet" in ct or "excel" in ct:
        return "excel"
    if "csv" in ct:
        return "csv"
    if "json" in ct:
        return "json"
    raise HTTPException(
        status_code=415,
        detail="Filformat stöds ej. Använd .xlsx, .csv eller .json.",
    )


def _rows_from_excel(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        result.append(dict(zip(headers, row)))
    return result


def _rows_from_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")  # Hanterar BOM
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _rows_from_json(content: bytes) -> list[dict]:
    data = json.loads(content.decode("utf-8"))
    if not isinstance(data, list):
        raise HTTPException(
            status_code=422, detail="JSON måste vara en array av objekt."
        )
    return data


def _coerce_row(row: dict, organization_id: UUID) -> dict:
    """Normalisera radvärden och injicera organization_id."""
    result = {}
    for key, val in row.items():
        # Tomma strängar → None
        if val == "" or val is None:
            result[key] = None
        else:
            result[key] = val

    result["organization_id"] = organization_id

    # bool-fält: hantera "true"/"false"/"1"/"0" från Excel/CSV
    bool_fields = [
        "nis2_applicable",
        "treats_personal_data",
        "treats_sensitive_data",
        "third_country_transfer",
        "has_elevated_protection",
        "security_protection",
        "dr_plan_exists",
    ]
    for field in bool_fields:
        val = result.get(field)
        if isinstance(val, str):
            result[field] = val.strip().lower() in ("true", "1", "yes", "ja")
        elif val is None:
            result[field] = False

    return result


async def _system_exists(
    db: AsyncSession, name: str, organization_id: UUID
) -> bool:
    stmt = select(System).where(
        System.name == name,
        System.organization_id == organization_id,
    )
    result = await db.execute(stmt)
    return result.scalar_one_or_none() is not None


async def _resolve_system(
    db: AsyncSession, name: str, organization_id: UUID | None = None
) -> System | None:
    """Slå upp system via namn (och optionellt org_id)."""
    stmt = select(System).where(System.name == name)
    if organization_id:
        stmt = stmt.where(System.organization_id == organization_id)
    result = await db.execute(stmt)
    return result.scalar_one_or_none()


@router.post("/systems")
async def import_systems(
    organization_id: UUID = Query(..., description="Organisation att importera till"),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_rls_db),
):
    """
    Importera system från Excel, CSV eller JSON.

    - Skippar rader där name+organization_id redan finns (ingen upsert).
    - Returnerar antal importerade + lista med valideringsfel per rad.
    """
    org = await db.get(Organization, organization_id)
    if not org:
        raise HTTPException(status_code=404, detail=f"Organisation {organization_id} finns inte")

    fmt = _detect_format(file.filename or "", file.content_type)
    content = await file.read()

    if fmt == "excel":
        rows = _rows_from_excel(content)
    elif fmt == "csv":
        rows = _rows_from_csv(content)
    else:
        rows = _rows_from_json(content)

    imported = 0
    errors = []

    for index, raw_row in enumerate(rows, start=2):  # rad 2 = första datarad (1 = rubrik)
        try:
            coerced = _coerce_row(raw_row, organization_id)
            data = SystemCreate.model_validate(coerced)
        except ValidationError as exc:
            errors.append({
                "row": index,
                "error": exc.errors(include_url=False),
            })
            continue
        except Exception as exc:
            errors.append({"row": index, "error": str(exc)})
            continue

        # Skippa om systemet redan finns
        exists = await _system_exists(db, data.name, organization_id)
        if exists:
            errors.append({
                "row": index,
                "error": f"System '{data.name}' finns redan i organisationen (skippad).",
            })
            continue

        system = System(**data.model_dump())
        db.add(system)
        imported += 1

    if imported > 0:
        await db.flush()

    return {"imported": imported, "errors": errors}


@router.post("/classifications")
async def import_classifications(
    organization_id: UUID | None = Query(None, description="Begränsa systemuppslagning till organisation"),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_rls_db),
):
    """
    Importera klassificeringar från Excel eller CSV.

    Kolumner: system_name, confidentiality (0-4), integrity (0-4),
    availability (0-4), traceability (0-4, valfri), classified_by.
    """
    fmt = _detect_format(file.filename or "", file.content_type)
    content = await file.read()

    if fmt == "excel":
        rows = _rows_from_excel(content)
    elif fmt == "csv":
        rows = _rows_from_csv(content)
    else:
        rows = _rows_from_json(content)

    imported = 0
    errors = []

    for index, raw_row in enumerate(rows, start=2):
        system_name = raw_row.get("system_name") or raw_row.get("System name")
        if not system_name:
            errors.append({"row": index, "error": "Saknar kolumn 'system_name'."})
            continue

        system = await _resolve_system(db, str(system_name).strip(), organization_id)
        if system is None:
            errors.append({
                "row": index,
                "error": f"System '{system_name}' hittades inte.",
            })
            continue

        try:
            confidentiality = int(raw_row.get("confidentiality") or 0)
            integrity = int(raw_row.get("integrity") or 0)
            availability = int(raw_row.get("availability") or 0)
            traceability_raw = raw_row.get("traceability")
            traceability = int(traceability_raw) if traceability_raw not in (None, "", "None") else None
            classified_by = raw_row.get("classified_by") or None
        except (ValueError, TypeError) as exc:
            errors.append({"row": index, "error": f"Ogiltigt värde: {exc}"})
            continue

        classification = SystemClassification(
            system_id=system.id,
            confidentiality=confidentiality,
            integrity=integrity,
            availability=availability,
            traceability=traceability,
            classified_by=classified_by,
        )
        db.add(classification)
        imported += 1

    if imported > 0:
        await db.flush()

    return {"imported": imported, "errors": errors}


_OWNER_ROLE_MAP = {role.value: role for role in OwnerRole}


@router.post("/owners")
async def import_owners(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_rls_db),
):
    """
    Importera systemägare från Excel eller CSV.

    Kolumner: system_name, name, email, phone, role, organization_id.
    Tillåtna roller: systemägare, informationsägare, systemförvaltare,
    teknisk_förvaltare, it_kontakt, dataskyddsombud.
    """
    fmt = _detect_format(file.filename or "", file.content_type)
    content = await file.read()

    if fmt == "excel":
        rows = _rows_from_excel(content)
    elif fmt == "csv":
        rows = _rows_from_csv(content)
    else:
        rows = _rows_from_json(content)

    imported = 0
    errors = []

    for index, raw_row in enumerate(rows, start=2):
        system_name = raw_row.get("system_name") or raw_row.get("System name")
        if not system_name:
            errors.append({"row": index, "error": "Saknar kolumn 'system_name'."})
            continue

        org_id_raw = raw_row.get("organization_id")
        try:
            row_org_id = UUID(str(org_id_raw).strip()) if org_id_raw not in (None, "", "None") else None
        except ValueError:
            errors.append({"row": index, "error": f"Ogiltigt organization_id: '{org_id_raw}'."})
            continue

        system = await _resolve_system(db, str(system_name).strip(), row_org_id)
        if system is None:
            errors.append({
                "row": index,
                "error": f"System '{system_name}' hittades inte.",
            })
            continue

        role_raw = str(raw_row.get("role") or "").strip().lower()
        role = _OWNER_ROLE_MAP.get(role_raw)
        if role is None:
            allowed = ", ".join(_OWNER_ROLE_MAP.keys())
            errors.append({
                "row": index,
                "error": f"Okänd roll '{role_raw}'. Tillåtna värden: {allowed}.",
            })
            continue

        owner = SystemOwner(
            system_id=system.id,
            name=raw_row.get("name") or None,
            email=raw_row.get("email") or None,
            phone=raw_row.get("phone") or None,
            role=role,
            organization_id=row_org_id,
        )
        db.add(owner)
        imported += 1

    if imported > 0:
        await db.flush()

    return {"imported": imported, "errors": errors}
