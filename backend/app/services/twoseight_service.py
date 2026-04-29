"""2C8 Excel-export-tjänst (Paket B.3).

2C8 stödjer manuell import av objekt och relationer via Excel. Detta
är "broadcast"-formatet tills ett 2C8 Extension Framework-plugin byggs
i Java/OSGi som en separat insats.
"""
from __future__ import annotations

import io
import zipfile
from uuid import UUID

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    BusinessCapability, BusinessProcess, BusinessRole, InformationAsset,
    OrgUnit, System, ValueStream,
    capability_system_link, process_system_link, process_capability_link,
    process_information_link, unit_capability_link,
)


_README = """\
2C8-import-paket — instruktioner

Detta paket innehåller ett Excel-blad med objekt och ett med relationer,
exporterat från Sundsvalls systemregister. Importera in i 2C8 Modeling Tool
i denna ordning:

1. Öppna ditt 2C8-projekt.
2. Verktyg → Import → Import från Excel.
3. Välj "objects.xlsx". Importera ett blad i taget eller alla på en gång —
   2C8 mappar 'id', 'namn', 'beskrivning' till Identitet/Namn/Beskrivning
   via standardobjekttyperna. Bekräfta kolumnmappningen vid behov.
4. Importera "relationships.xlsx" på samma sätt. 2C8 matchar käll-/mål-id
   mot redan importerade objekt.
5. Vid behov: skapa en vy som visar objekten — datat ligger nu i 2C8.

Vid uppdatering — kör om importen från registret. Dataägarskapet ligger
i registret, visualiseringen i 2C8.

Frågor: Håkan Simonsson, DigIT (hakan.simonsson@sundsvall.se)
"""


def _id(uuid_val: UUID) -> str:
    return str(uuid_val)


async def _objects_workbook(
    organization_id: UUID, db: AsyncSession,
) -> Workbook:
    wb = Workbook()
    # Default-bladet kommer användas som första
    default = wb.active
    if default is not None:
        wb.remove(default)

    # Förmåga
    caps = (await db.execute(
        select(BusinessCapability)
        .where(BusinessCapability.organization_id == organization_id)
        .order_by(BusinessCapability.name)
    )).scalars().all()
    sheet = wb.create_sheet("Förmåga")
    sheet.append(["id", "namn", "beskrivning", "parent_id", "ägare", "mognad"])
    for c in caps:
        sheet.append([
            _id(c.id), c.name, c.description or "",
            _id(c.parent_capability_id) if c.parent_capability_id else "",
            c.capability_owner or "",
            c.maturity_level if c.maturity_level is not None else "",
        ])

    # Process
    processes = (await db.execute(
        select(BusinessProcess)
        .where(BusinessProcess.organization_id == organization_id)
        .order_by(BusinessProcess.name)
    )).scalars().all()
    sheet = wb.create_sheet("Process")
    sheet.append(["id", "namn", "beskrivning", "parent_id", "ägare", "kritikalitet"])
    for p in processes:
        sheet.append([
            _id(p.id), p.name, p.description or "",
            _id(p.parent_process_id) if p.parent_process_id else "",
            p.process_owner or "",
            p.criticality.value if p.criticality else "",
        ])

    # Värdeström
    streams = (await db.execute(
        select(ValueStream)
        .where(ValueStream.organization_id == organization_id)
        .order_by(ValueStream.name)
    )).scalars().all()
    sheet = wb.create_sheet("Värdeström")
    sheet.append(["id", "namn", "beskrivning", "etapper"])
    for v in streams:
        stages = ", ".join(s.get("name", "") for s in (v.stages or []))
        sheet.append([_id(v.id), v.name, v.description or "", stages])

    # Organisationsenhet
    units = (await db.execute(
        select(OrgUnit)
        .where(OrgUnit.organization_id == organization_id)
        .order_by(OrgUnit.name)
    )).scalars().all()
    sheet = wb.create_sheet("Organisationsenhet")
    sheet.append(["id", "namn", "parent_id", "typ", "chef", "kostnadsställe"])
    for u in units:
        sheet.append([
            _id(u.id), u.name,
            _id(u.parent_unit_id) if u.parent_unit_id else "",
            u.unit_type.value, u.manager_name or "", u.cost_center or "",
        ])

    # System (Applikation)
    systems = (await db.execute(
        select(System)
        .where(System.organization_id == organization_id)
        .order_by(System.name)
    )).scalars().all()
    sheet = wb.create_sheet("System (Applikation)")
    sheet.append(["id", "namn", "kategori", "kritikalitet", "livscykel"])
    for s in systems:
        sheet.append([
            _id(s.id), s.name, s.system_category.value,
            s.criticality.value if s.criticality else "",
            s.lifecycle_status.value if s.lifecycle_status else "",
        ])

    # Informationsmängd
    assets = (await db.execute(
        select(InformationAsset)
        .where(InformationAsset.organization_id == organization_id)
        .order_by(InformationAsset.name)
    )).scalars().all()
    sheet = wb.create_sheet("Informationsmängd")
    sheet.append(["id", "namn", "K", "R", "T", "gallring"])
    for a in assets:
        sheet.append([
            _id(a.id), a.name,
            a.confidentiality if a.confidentiality is not None else "",
            a.integrity if a.integrity is not None else "",
            a.availability if a.availability is not None else "",
            a.retention_period or "",
        ])

    # Roll
    roles = (await db.execute(
        select(BusinessRole)
        .where(BusinessRole.organization_id == organization_id)
        .order_by(BusinessRole.name)
    )).scalars().all()
    sheet = wb.create_sheet("Roll")
    sheet.append(["id", "namn", "beskrivning", "ägare"])
    for r in roles:
        sheet.append([_id(r.id), r.name, r.description or "", r.role_owner or ""])

    return wb


async def _relationships_workbook(
    organization_id: UUID, db: AsyncSession,
) -> Workbook:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Relationer"
    sheet.append(["källa_typ", "källa_id", "relations_typ", "mål_typ", "mål_id"])

    # Filter — bara relationer där minst ena änden tillhör organisationen
    sys_ids: set[UUID] = {
        s.id for s in (await db.execute(
            select(System.id).where(System.organization_id == organization_id)
        )).all()
    }
    cap_ids: set[UUID] = {
        c.id for c in (await db.execute(
            select(BusinessCapability.id)
            .where(BusinessCapability.organization_id == organization_id)
        )).all()
    }
    proc_ids: set[UUID] = {
        p.id for p in (await db.execute(
            select(BusinessProcess.id)
            .where(BusinessProcess.organization_id == organization_id)
        )).all()
    }
    info_ids: set[UUID] = {
        a.id for a in (await db.execute(
            select(InformationAsset.id)
            .where(InformationAsset.organization_id == organization_id)
        )).all()
    }
    unit_ids: set[UUID] = {
        u.id for u in (await db.execute(
            select(OrgUnit.id).where(OrgUnit.organization_id == organization_id)
        )).all()
    }

    # Composition (hierarki)
    caps = (await db.execute(
        select(BusinessCapability)
        .where(BusinessCapability.organization_id == organization_id)
    )).scalars().all()
    for c in caps:
        if c.parent_capability_id:
            sheet.append([
                "Förmåga", _id(c.parent_capability_id),
                "Består av",
                "Förmåga", _id(c.id),
            ])

    procs = (await db.execute(
        select(BusinessProcess)
        .where(BusinessProcess.organization_id == organization_id)
    )).scalars().all()
    for p in procs:
        if p.parent_process_id:
            sheet.append([
                "Process", _id(p.parent_process_id),
                "Består av",
                "Process", _id(p.id),
            ])

    units = (await db.execute(
        select(OrgUnit)
        .where(OrgUnit.organization_id == organization_id)
    )).scalars().all()
    for u in units:
        if u.parent_unit_id:
            sheet.append([
                "Organisationsenhet", _id(u.parent_unit_id),
                "Består av",
                "Organisationsenhet", _id(u.id),
            ])

    # Realiserar — System → Förmåga
    cap_sys = (await db.execute(
        select(capability_system_link.c.capability_id, capability_system_link.c.system_id)
    )).all()
    for cap_id, sys_id in cap_sys:
        if cap_id in cap_ids and sys_id in sys_ids:
            sheet.append([
                "System (Applikation)", _id(sys_id),
                "Realiserar",
                "Förmåga", _id(cap_id),
            ])

    # Stödjer — System → Process
    proc_sys = (await db.execute(
        select(process_system_link.c.process_id, process_system_link.c.system_id)
    )).all()
    for proc_id, sys_id in proc_sys:
        if proc_id in proc_ids and sys_id in sys_ids:
            sheet.append([
                "System (Applikation)", _id(sys_id),
                "Stödjer",
                "Process", _id(proc_id),
            ])

    # Realiserar — Process → Förmåga
    proc_cap = (await db.execute(
        select(process_capability_link.c.process_id, process_capability_link.c.capability_id)
    )).all()
    for proc_id, cap_id in proc_cap:
        if proc_id in proc_ids and cap_id in cap_ids:
            sheet.append([
                "Process", _id(proc_id),
                "Realiserar",
                "Förmåga", _id(cap_id),
            ])

    # Använder — Process → Informationsmängd
    proc_info = (await db.execute(
        select(
            process_information_link.c.process_id,
            process_information_link.c.information_asset_id,
        )
    )).all()
    for proc_id, info_id in proc_info:
        if proc_id in proc_ids and info_id in info_ids:
            sheet.append([
                "Process", _id(proc_id),
                "Använder",
                "Informationsmängd", _id(info_id),
            ])

    # Tillhör — Förmåga → Organisationsenhet
    unit_cap = (await db.execute(
        select(unit_capability_link.c.unit_id, unit_capability_link.c.capability_id)
    )).all()
    for unit_id, cap_id in unit_cap:
        if unit_id in unit_ids and cap_id in cap_ids:
            sheet.append([
                "Förmåga", _id(cap_id),
                "Tillhör",
                "Organisationsenhet", _id(unit_id),
            ])

    return wb


def _wb_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def build_objects_xlsx(organization_id: UUID, db: AsyncSession) -> bytes:
    return _wb_bytes(await _objects_workbook(organization_id, db))


async def build_relationships_xlsx(organization_id: UUID, db: AsyncSession) -> bytes:
    return _wb_bytes(await _relationships_workbook(organization_id, db))


async def build_full_package_zip(organization_id: UUID, db: AsyncSession) -> bytes:
    objects = await build_objects_xlsx(organization_id, db)
    relationships = await build_relationships_xlsx(organization_id, db)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("objects.xlsx", objects)
        zf.writestr("relationships.xlsx", relationships)
        zf.writestr("README.txt", _README)
    return buf.getvalue()
