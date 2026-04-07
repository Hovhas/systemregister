"""Business logic for file import (Excel, CSV, JSON)."""

import csv
import io
import json
from uuid import UUID

from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import System
from app.models.enums import OwnerRole
from app.models.models import Organization, SystemClassification, SystemOwner
from app.schemas import SystemCreate

SUPPORTED_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/csv",
    "application/csv",
    "application/json",
    "text/plain",
}


class ImportService:

    @staticmethod
    def detect_format(filename: str, content_type: str | None) -> str:
        name = (filename or "").lower()
        if name.endswith(".xlsx") or name.endswith(".xls"):
            return "excel"
        if name.endswith(".csv"):
            return "csv"
        if name.endswith(".json"):
            return "json"
        ct = (content_type or "").lower()
        if "spreadsheet" in ct or "excel" in ct:
            return "excel"
        if "csv" in ct:
            return "csv"
        if "json" in ct:
            return "json"
        raise ValueError("Filformat stöds ej. Använd .xlsx, .csv eller .json.")

    @staticmethod
    def rows_from_excel(content: bytes) -> list[dict]:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]

    @staticmethod
    def rows_from_csv(content: bytes) -> list[dict]:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    @staticmethod
    def rows_from_json(content: bytes) -> list[dict]:
        try:
            data = json.loads(content.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError) as exc:
            raise ValueError(f"Ogiltig JSON-syntax: {exc}")
        if not isinstance(data, list):
            raise ValueError("JSON måste vara en array av objekt.")
        return data

    @staticmethod
    def parse_rows(content: bytes, fmt: str) -> list[dict]:
        if fmt == "excel":
            return ImportService.rows_from_excel(content)
        elif fmt == "csv":
            return ImportService.rows_from_csv(content)
        else:
            return ImportService.rows_from_json(content)

    @staticmethod
    def coerce_system_row(row: dict, organization_id: UUID) -> dict:
        result = {}
        for key, val in row.items():
            result[key] = None if val == "" or val is None else val
        result["organization_id"] = organization_id
        bool_fields = [
            "nis2_applicable", "treats_personal_data", "treats_sensitive_data",
            "third_country_transfer", "has_elevated_protection", "security_protection",
            "dr_plan_exists",
        ]
        for field in bool_fields:
            val = result.get(field)
            if isinstance(val, str):
                result[field] = val.strip().lower() in ("true", "1", "yes", "ja")
            elif val is None:
                result[field] = False
        return result

    @staticmethod
    async def system_exists(db: AsyncSession, name: str, organization_id: UUID) -> bool:
        stmt = select(System).where(System.name == name, System.organization_id == organization_id)
        result = await db.execute(stmt)
        return result.scalar_one_or_none() is not None

    @staticmethod
    async def resolve_system(db: AsyncSession, name: str, organization_id: UUID | None = None) -> System | None:
        stmt = select(System).where(System.name == name)
        if organization_id:
            stmt = stmt.where(System.organization_id == organization_id)
        result = await db.execute(stmt)
        return result.scalar_one_or_none()

    @staticmethod
    async def import_systems(
        db: AsyncSession, rows: list[dict], organization_id: UUID,
    ) -> tuple[int, list[dict]]:
        imported = 0
        errors = []
        for index, raw_row in enumerate(rows, start=2):
            try:
                coerced = ImportService.coerce_system_row(raw_row, organization_id)
                data = SystemCreate.model_validate(coerced)
            except ValidationError as exc:
                errors.append({"row": index, "error": exc.errors(include_url=False)})
                continue
            except Exception as exc:
                errors.append({"row": index, "error": str(exc)})
                continue

            if await ImportService.system_exists(db, data.name, organization_id):
                errors.append({"row": index, "error": f"System '{data.name}' finns redan i organisationen (skippad)."})
                continue

            db.add(System(**data.model_dump()))
            imported += 1

        if imported > 0:
            await db.flush()
        return imported, errors

    @staticmethod
    async def import_classifications(
        db: AsyncSession, rows: list[dict], organization_id: UUID | None,
    ) -> tuple[int, list[dict]]:
        imported = 0
        errors = []
        for index, raw_row in enumerate(rows, start=2):
            system_name = raw_row.get("system_name") or raw_row.get("System name")
            if not system_name:
                errors.append({"row": index, "error": "Saknar kolumn 'system_name'."})
                continue
            system = await ImportService.resolve_system(db, str(system_name).strip(), organization_id)
            if system is None:
                errors.append({"row": index, "error": f"System '{system_name}' hittades inte."})
                continue
            try:
                confidentiality = int(raw_row.get("confidentiality") or 0)
                integrity = int(raw_row.get("integrity") or 0)
                availability = int(raw_row.get("availability") or 0)
                traceability_raw = raw_row.get("traceability")
                traceability = int(traceability_raw) if traceability_raw not in (None, "", "None") else None
                classified_by = raw_row.get("classified_by") or None
            except (ValueError, TypeError) as exc:
                errors.append({"row": index, "error": f"Ogiltigt värde: {exc}"})
                continue
            db.add(SystemClassification(
                system_id=system.id,
                confidentiality=confidentiality, integrity=integrity,
                availability=availability, traceability=traceability,
                classified_by=classified_by,
            ))
            imported += 1
        if imported > 0:
            await db.flush()
        return imported, errors

    OWNER_ROLE_MAP = {role.value: role for role in OwnerRole}

    @staticmethod
    async def import_owners(
        db: AsyncSession, rows: list[dict],
    ) -> tuple[int, list[dict]]:
        imported = 0
        errors = []
        for index, raw_row in enumerate(rows, start=2):
            system_name = raw_row.get("system_name") or raw_row.get("System name")
            if not system_name:
                errors.append({"row": index, "error": "Saknar kolumn 'system_name'."})
                continue
            org_id_raw = raw_row.get("organization_id")
            try:
                row_org_id = UUID(str(org_id_raw).strip()) if org_id_raw not in (None, "", "None") else None
            except ValueError:
                errors.append({"row": index, "error": f"Ogiltigt organization_id: '{org_id_raw}'."})
                continue
            system = await ImportService.resolve_system(db, str(system_name).strip(), row_org_id)
            if system is None:
                errors.append({"row": index, "error": f"System '{system_name}' hittades inte."})
                continue
            role_raw = str(raw_row.get("role") or "").strip().lower()
            role = ImportService.OWNER_ROLE_MAP.get(role_raw)
            if role is None:
                allowed = ", ".join(ImportService.OWNER_ROLE_MAP.keys())
                errors.append({"row": index, "error": f"Okänd roll '{role_raw}'. Tillåtna värden: {allowed}."})
                continue
            owner_name = (raw_row.get("name") or "").strip()
            if not owner_name:
                errors.append({"row": index, "error": "Saknar kolumn 'name' eller värdet är tomt."})
                continue
            db.add(SystemOwner(
                system_id=system.id, name=owner_name,
                email=raw_row.get("email") or None, phone=raw_row.get("phone") or None,
                role=role, organization_id=row_org_id,
            ))
            imported += 1
        if imported > 0:
            await db.flush()
        return imported, errors
