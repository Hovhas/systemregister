"""
Tests for /api/v1/export endpoints (JSON, CSV, XLSX).
"""

import csv
import io

import pytest
from uuid import uuid4

from tests.factories import create_org, create_system


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_json(client):
    """GET /api/v1/export/systems.json returns a valid JSON array of systems."""
    org = await create_org(client)
    await create_system(client, org["id"], name="Procapita")
    await create_system(client, org["id"], name="Visma Lön")

    resp = await client.get("/api/v1/export/systems.json")

    assert resp.status_code == 200, f"Expected 200: {resp.text}"

    # Verify Content-Disposition header
    content_disposition = resp.headers.get("content-disposition", "")
    assert "systems.json" in content_disposition, (
        f"Expected filename in Content-Disposition, got: {content_disposition}"
    )

    # Verify parseable JSON array
    data = resp.json()
    assert isinstance(data, list), "Response should be a JSON array"
    assert len(data) >= 2, f"Expected at least 2 systems, got {len(data)}"

    # Verify expected fields are present
    first = data[0]
    assert "name" in first
    assert "system_category" in first
    assert "criticality" in first
    assert "lifecycle_status" in first


@pytest.mark.asyncio
async def test_export_json_empty(client):
    """GET /api/v1/export/systems.json with no systems returns empty array."""
    resp = await client.get("/api/v1/export/systems.json")

    assert resp.status_code == 200
    data = resp.json()
    assert data == [], f"Expected empty array, got {data}"


@pytest.mark.asyncio
async def test_export_csv(client):
    """GET /api/v1/export/systems.csv returns valid CSV with headers and data rows."""
    org = await create_org(client)
    await create_system(client, org["id"], name="Procapita")
    await create_system(client, org["id"], name="Pulsen Combine")

    resp = await client.get("/api/v1/export/systems.csv")

    assert resp.status_code == 200, f"Expected 200: {resp.text}"

    # Verify Content-Type
    content_type = resp.headers.get("content-type", "")
    assert "text/csv" in content_type, f"Expected text/csv, got: {content_type}"

    # Verify Content-Disposition
    content_disposition = resp.headers.get("content-disposition", "")
    assert "systems.csv" in content_disposition, (
        f"Expected filename in Content-Disposition, got: {content_disposition}"
    )

    # Parse and validate CSV
    text = resp.text
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)

    assert len(rows) >= 2, f"Expected at least 2 data rows, got {len(rows)}"

    # Header should include expected columns
    fieldnames = reader.fieldnames or []
    assert "name" in fieldnames, "CSV should have 'name' column"
    assert "system_category" in fieldnames, "CSV should have 'system_category' column"
    assert "criticality" in fieldnames, "CSV should have 'criticality' column"

    # Data values should be populated
    names = [row["name"] for row in rows]
    assert "Procapita" in names or "Pulsen Combine" in names, (
        f"Expected created systems in export, got names: {names}"
    )


@pytest.mark.asyncio
async def test_export_csv_empty(client):
    """GET /api/v1/export/systems.csv with no systems returns only header row."""
    resp = await client.get("/api/v1/export/systems.csv")

    assert resp.status_code == 200
    content_type = resp.headers.get("content-type", "")
    assert "text/csv" in content_type

    reader = csv.DictReader(io.StringIO(resp.text))
    rows = list(reader)
    assert rows == [], "Expected no data rows in empty export"


@pytest.mark.asyncio
async def test_export_xlsx(client):
    """GET /api/v1/export/systems.xlsx returns valid XLSX with correct Content-Type."""
    org = await create_org(client)
    await create_system(client, org["id"], name="Procapita")

    resp = await client.get("/api/v1/export/systems.xlsx")

    assert resp.status_code == 200, f"Expected 200: {resp.text}"

    # Verify Content-Type for Excel
    content_type = resp.headers.get("content-type", "")
    assert "spreadsheetml.sheet" in content_type, (
        f"Expected OOXML content-type, got: {content_type}"
    )

    # Verify Content-Disposition
    content_disposition = resp.headers.get("content-disposition", "")
    assert "systems.xlsx" in content_disposition, (
        f"Expected xlsx filename in Content-Disposition, got: {content_disposition}"
    )

    # Verify the bytes are a valid XLSX (ZIP magic bytes: PK\x03\x04)
    content = resp.content
    assert content[:4] == b"PK\x03\x04", "XLSX file should start with ZIP magic bytes"


@pytest.mark.asyncio
async def test_export_xlsx_parseable(client):
    """GET /api/v1/export/systems.xlsx returns a file parseable by openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    org = await create_org(client)
    await create_system(client, org["id"], name="Procapita")
    await create_system(client, org["id"], name="Visma Lön")

    resp = await client.get("/api/v1/export/systems.xlsx")
    assert resp.status_code == 200

    wb = load_workbook(filename=io.BytesIO(resp.content))
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    # First row is header, subsequent rows are data
    assert len(rows) >= 3, f"Expected header + at least 2 data rows, got {len(rows)}"
    header = rows[0]
    assert "name" in header, "Header row should contain 'name'"


@pytest.mark.asyncio
async def test_export_filtered_by_org(client):
    """Export with organization_id filter returns only that org's systems."""
    org1 = await create_org(client, name="Org 1", org_number="111111-1111")
    org2 = await create_org(client, name="Org 2", org_number="222222-2222")

    await create_system(client, org1["id"], name="Org1 System")
    await create_system(client, org2["id"], name="Org2 System")

    # Export only org1's systems as JSON
    resp = await client.get("/api/v1/export/systems.json", params={"organization_id": org1["id"]})

    assert resp.status_code == 200
    data = resp.json()
    assert len(data) == 1, f"Expected 1 system for org1, got {len(data)}"
    assert data[0]["name"] == "Org1 System", f"Expected 'Org1 System', got {data[0]['name']}"


@pytest.mark.asyncio
async def test_export_filtered_by_org_csv(client):
    """CSV export with organization_id filter returns only that org's systems."""
    org1 = await create_org(client, name="Filter Org A", org_number="333333-3333")
    org2 = await create_org(client, name="Filter Org B", org_number="444444-4444")

    await create_system(client, org1["id"], name="Org A System")
    await create_system(client, org2["id"], name="Org B System")

    resp = await client.get("/api/v1/export/systems.csv", params={"organization_id": org1["id"]})

    assert resp.status_code == 200
    reader = csv.DictReader(io.StringIO(resp.text))
    rows = list(reader)
    assert len(rows) == 1, f"Expected 1 row for org A, got {len(rows)}"
    assert rows[0]["name"] == "Org A System"

    # Verify the filename includes the org id
    content_disposition = resp.headers.get("content-disposition", "")
    assert org1["id"] in content_disposition, (
        f"Filtered export filename should include org id, got: {content_disposition}"
    )


@pytest.mark.asyncio
async def test_export_sorted_by_name(client):
    """Exported JSON should be sorted alphabetically by name."""
    org = await create_org(client)
    await create_system(client, org["id"], name="Zebra System")
    await create_system(client, org["id"], name="Alpha System")
    await create_system(client, org["id"], name="Medel System")

    resp = await client.get("/api/v1/export/systems.json")

    assert resp.status_code == 200
    data = resp.json()
    names = [row["name"] for row in data]
    assert names == sorted(names), f"Export should be sorted by name, got: {names}"


# ---------------------------------------------------------------------------
# Utökade exporttester — Kategori 8
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_csv_parses_all_rows(client):
    """CSV-export innehåller en rad per system, inga extra rader."""
    org = await create_org(client, name="CSVParseOrg", org_number="111222-3333")
    for i in range(3):
        await create_system(client, org["id"], name=f"CsvRow-{i}-{uuid4().hex[:6]}")

    resp = await client.get("/api/v1/export/systems.csv", params={"organization_id": org["id"]})
    assert resp.status_code == 200

    reader = csv.DictReader(io.StringIO(resp.text))
    rows = list(reader)
    assert len(rows) == 3, f"Förväntade 3 rader, fick {len(rows)}"


@pytest.mark.asyncio
async def test_export_csv_contains_expected_columns(client):
    """CSV-exporten innehåller alla förväntade kolumner."""
    org = await create_org(client, name="ColCheckOrg", org_number="555666-7777")
    await create_system(client, org["id"], name="ColCheckSys")

    resp = await client.get("/api/v1/export/systems.csv")
    assert resp.status_code == 200

    reader = csv.DictReader(io.StringIO(resp.text))
    list(reader)  # consume rows
    fieldnames = reader.fieldnames or []

    expected_cols = ["name", "system_category", "criticality", "lifecycle_status"]
    for col in expected_cols:
        assert col in fieldnames, f"Kolumn '{col}' saknas i CSV-exporten"


@pytest.mark.asyncio
async def test_export_json_contains_required_fields(client):
    """JSON-exporterade objekt innehåller alla nödvändiga fält."""
    org = await create_org(client, name="JSONFieldOrg", org_number="888999-0000")
    await create_system(client, org["id"], name="JSONFieldSys")

    resp = await client.get("/api/v1/export/systems.json")
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) >= 1

    first = data[0]
    required_fields = ["id", "name", "system_category", "criticality",
                       "lifecycle_status", "organization_id", "created_at"]
    for field in required_fields:
        assert field in first, f"Fält '{field}' saknas i JSON-exporten"


@pytest.mark.asyncio
async def test_export_json_nis2_fields_present(client):
    """JSON-export innehåller NIS2-relaterade fält."""
    org = await create_org(client, name="NIS2ExportOrg", org_number="100200-3004")
    await create_system(client, org["id"], name="NIS2ExportSys")

    resp = await client.get("/api/v1/export/systems.json")
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) >= 1

    first = data[0]
    nis2_fields = ["nis2_applicable", "nis2_classification"]
    for field in nis2_fields:
        assert field in first, f"NIS2-fält '{field}' saknas i JSON-exporten"


@pytest.mark.asyncio
async def test_export_json_gdpr_fields_present(client):
    """JSON-export innehåller GDPR-relaterade flaggor."""
    org = await create_org(client, name="GDPRExportOrg", org_number="100200-3005")
    await create_system(client, org["id"], name="GDPRExportSys")

    resp = await client.get("/api/v1/export/systems.json")
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) >= 1

    first = data[0]
    gdpr_fields = ["treats_personal_data", "treats_sensitive_data"]
    for field in gdpr_fields:
        assert field in first, f"GDPR-fält '{field}' saknas i JSON-exporten"


@pytest.mark.asyncio
async def test_export_xlsx_org_filter(client):
    """XLSX-export med organization_id filter returnerar enbart org:s system."""
    org1 = await create_org(client, name="XLSXOrgA", org_number="XA1-001")
    org2 = await create_org(client, name="XLSXOrgB", org_number="XB2-002")

    await create_system(client, org1["id"], name="XLSX Org1 Sys")
    await create_system(client, org2["id"], name="XLSX Org2 Sys")

    resp = await client.get("/api/v1/export/systems.xlsx",
                             params={"organization_id": org1["id"]})
    assert resp.status_code == 200

    try:
        from openpyxl import load_workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    wb = load_workbook(filename=io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    assert len(rows) >= 2, "Skall ha minst header + 1 datarad"
    # Verifiera att enbart Org1:s system finns
    header = list(rows[0])
    name_idx = header.index("name") if "name" in header else None
    if name_idx is not None:
        data_names = [row[name_idx] for row in rows[1:] if row[name_idx]]
        assert "XLSX Org1 Sys" in data_names
        assert "XLSX Org2 Sys" not in data_names


@pytest.mark.asyncio
async def test_export_xlsx_empty_database(client):
    """XLSX-export med tom databas returnerar fil med enbart header-rad."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    resp = await client.get("/api/v1/export/systems.xlsx")
    assert resp.status_code == 200

    wb = load_workbook(filename=io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Skall ha header-rad men inga datarader
    assert len(rows) >= 1, "XLSX skall alltid ha header-rad"


@pytest.mark.asyncio
async def test_export_json_invalid_org_filter_returns_empty(client):
    """JSON-export med okänt organization_id returnerar tom lista."""
    fake_org = "00000000-0000-0000-0000-000000000000"
    resp = await client.get("/api/v1/export/systems.json",
                             params={"organization_id": fake_org})
    assert resp.status_code == 200
    assert resp.json() == [], "Okänt organization_id borde ge tom lista"


@pytest.mark.asyncio
async def test_export_csv_invalid_org_filter_returns_header_only(client):
    """CSV-export med okänt organization_id returnerar enbart header-rad."""
    fake_org = "00000000-0000-0000-0000-000000000000"
    resp = await client.get("/api/v1/export/systems.csv",
                             params={"organization_id": fake_org})
    assert resp.status_code == 200

    reader = csv.DictReader(io.StringIO(resp.text))
    rows = list(reader)
    assert rows == [], "Okänt org_id i CSV-filter borde ge inga rader"


@pytest.mark.asyncio
async def test_export_json_large_dataset(client):
    """JSON-export hanterar dataset med 50 system."""
    org = await create_org(client, name="LargeExportOrg", org_number="LRG-001")
    for i in range(50):
        await create_system(client, org["id"], name=f"LgExp-{i:03d}-{uuid4().hex[:6]}")

    resp = await client.get("/api/v1/export/systems.json",
                             params={"organization_id": org["id"]})
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) == 50, f"Förväntade 50 system, fick {len(data)}"


@pytest.mark.asyncio
async def test_export_csv_large_dataset(client):
    """CSV-export hanterar dataset med 50 system."""
    org = await create_org(client, name="LargeCSVOrg", org_number="LCV-002")
    for i in range(50):
        await create_system(client, org["id"], name=f"LgCsv-{i:03d}-{uuid4().hex[:6]}")

    resp = await client.get("/api/v1/export/systems.csv",
                             params={"organization_id": org["id"]})
    assert resp.status_code == 200

    reader = csv.DictReader(io.StringIO(resp.text))
    rows = list(reader)
    assert len(rows) == 50, f"Förväntade 50 rader i CSV, fick {len(rows)}"


@pytest.mark.asyncio
async def test_export_json_system_with_all_flags(client):
    """JSON-export av system med alla flaggor satta exporterar korrekt."""
    org = await create_org(client, name="FlagExportOrg", org_number="FLAG-001")
    await create_system(client, org["id"], name="FlagExportSys",
                        criticality="kritisk",
                        lifecycle_status="i_drift",
                        nis2_applicable=True,
                        nis2_classification="väsentlig",
                        treats_personal_data=True,
                        treats_sensitive_data=True,
                        third_country_transfer=True,
                        has_elevated_protection=True)

    resp = await client.get("/api/v1/export/systems.json",
                             params={"organization_id": org["id"]})
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) == 1

    sys = data[0]
    assert sys["criticality"] == "kritisk"
    assert sys["nis2_applicable"] is True
    assert sys["treats_personal_data"] is True


@pytest.mark.asyncio
async def test_export_content_type_json(client):
    """JSON-export har korrekt Content-Type."""
    resp = await client.get("/api/v1/export/systems.json")
    assert resp.status_code == 200
    content_type = resp.headers.get("content-type", "")
    assert "json" in content_type, f"Förväntade JSON content-type, fick: {content_type}"


@pytest.mark.asyncio
async def test_export_xlsx_content_type(client):
    """XLSX-export har korrekt Content-Type."""
    resp = await client.get("/api/v1/export/systems.xlsx")
    assert resp.status_code == 200
    content_type = resp.headers.get("content-type", "")
    assert "spreadsheetml" in content_type or "openxmlformats" in content_type, (
        f"Förväntade XLSX content-type, fick: {content_type}"
    )


@pytest.mark.asyncio
async def test_export_csv_unicode_in_names(client):
    """CSV-export hanterar svenska tecken och specialtecken korrekt."""
    org = await create_org(client, name="ÅÄÖ Org", org_number="AAO-001")
    await create_system(client, org["id"], name="Åäö System med specialtecken éàü",
                        description="Beskrivning med specialtecken")

    resp = await client.get("/api/v1/export/systems.csv",
                             params={"organization_id": org["id"]})
    assert resp.status_code == 200

    reader = csv.DictReader(io.StringIO(resp.text))
    rows = list(reader)
    assert len(rows) == 1
    assert "Åäö" in rows[0]["name"]


@pytest.mark.asyncio
async def test_export_json_multi_org_no_filter_returns_all(client):
    """JSON-export utan org-filter returnerar system från alla orgs."""
    org_a = await create_org(client, name="MultiExportA", org_number="MEA-001")
    org_b = await create_org(client, name="MultiExportB", org_number="MEB-002")

    sys_a = await create_system(client, org_a["id"], name="MultiExportSysA")
    sys_b = await create_system(client, org_b["id"], name="MultiExportSysB")

    resp = await client.get("/api/v1/export/systems.json")
    assert resp.status_code == 200
    ids = [s["id"] for s in resp.json()]
    assert sys_a["id"] in ids
    assert sys_b["id"] in ids
