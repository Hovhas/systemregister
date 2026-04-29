"""Tester för 2C8-export (Paket B.3)."""
import io
import zipfile

import pytest
from openpyxl import load_workbook

from tests.factories import create_org, create_system


async def _create_capability(client, org_id: str, name: str, **extra):
    resp = await client.post(
        "/api/v1/capabilities/",
        json={"organization_id": str(org_id), "name": name, **extra},
    )
    assert resp.status_code == 201
    return resp.json()


@pytest.mark.asyncio
async def test_objects_xlsx_has_expected_sheets(client):
    org = await create_org(client)
    await create_system(client, org["id"], name="System1")
    await _create_capability(client, org["id"], "Förmåga1")

    resp = await client.get(
        "/api/v1/export/2c8/objects.xlsx",
        params={"organization_id": org["id"]},
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    expected = {
        "Förmåga", "Process", "Värdeström", "Organisationsenhet",
        "System (Applikation)", "Informationsmängd", "Roll",
    }
    assert expected <= set(wb.sheetnames)

    sys_sheet = wb["System (Applikation)"]
    headers = [c.value for c in sys_sheet[1]]
    assert headers[:5] == ["id", "namn", "kategori", "kritikalitet", "livscykel"]
    # Minst en datarad utöver header
    rows = list(sys_sheet.iter_rows(values_only=True))
    assert len(rows) >= 2
    assert any(r[1] == "System1" for r in rows[1:])


@pytest.mark.asyncio
async def test_relationships_xlsx_has_expected_columns(client):
    org = await create_org(client)
    sys_a = await create_system(client, org["id"], name="System2")
    cap = await _create_capability(client, org["id"], "Förmåga2")
    await client.post(
        f"/api/v1/capabilities/{cap['id']}/systems",
        json={"system_id": sys_a["id"]},
    )

    resp = await client.get(
        "/api/v1/export/2c8/relationships.xlsx",
        params={"organization_id": org["id"]},
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    sheet = wb["Relationer"]
    headers = [c.value for c in sheet[1]]
    assert headers == ["källa_typ", "källa_id", "relations_typ", "mål_typ", "mål_id"]
    rows = list(sheet.iter_rows(values_only=True))[1:]
    assert any(r[2] == "Realiserar" for r in rows)


@pytest.mark.asyncio
async def test_full_package_zip(client):
    org = await create_org(client)
    await create_system(client, org["id"], name="System3")

    resp = await client.get(
        "/api/v1/export/2c8/full-package.zip",
        params={"organization_id": org["id"]},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/zip"

    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        names = set(zf.namelist())
        assert {"objects.xlsx", "relationships.xlsx", "README.txt"} <= names
        readme = zf.read("README.txt").decode("utf-8")
        assert "2C8" in readme


@pytest.mark.asyncio
async def test_relationships_reference_existing_object_ids(client):
    """Alla källid/målid i relations-bladet ska ha en motsvarande rad i objects."""
    org = await create_org(client)
    sys_a = await create_system(client, org["id"], name="SysA")
    cap = await _create_capability(client, org["id"], "Förm")
    await client.post(
        f"/api/v1/capabilities/{cap['id']}/systems",
        json={"system_id": sys_a["id"]},
    )

    obj_resp = await client.get(
        "/api/v1/export/2c8/objects.xlsx",
        params={"organization_id": org["id"]},
    )
    rel_resp = await client.get(
        "/api/v1/export/2c8/relationships.xlsx",
        params={"organization_id": org["id"]},
    )
    obj_wb = load_workbook(io.BytesIO(obj_resp.content))
    rel_wb = load_workbook(io.BytesIO(rel_resp.content))

    all_ids: set[str] = set()
    for sheet_name in obj_wb.sheetnames:
        sheet = obj_wb[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                all_ids.add(str(row[0]))

    rel_sheet = rel_wb["Relationer"]
    for row in rel_sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        src_id, tgt_id = str(row[1]), str(row[4])
        assert src_id in all_ids, f"källa-id {src_id} saknas i objects"
        assert tgt_id in all_ids, f"mål-id {tgt_id} saknas i objects"
